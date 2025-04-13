from fastapi import APIRouter, UploadFile, File, Depends, HTTPException
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, text
from openpyxl import load_workbook
from datetime import datetime, date
from decimal import Decimal

from app.core.database import get_db
from app.models import Plan, Dictionary

router = APIRouter()


@router.post("/plans_insert")
async def upload_plans(file: UploadFile = File(...), db: AsyncSession = Depends(get_db)):
    if not file.filename.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Invalid file format. Expected .xlsx")

    workbook = load_workbook(filename=file.file, data_only=True)
    sheet = workbook.active

    errors = []
    plans_to_insert = []

    for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if not row or all(cell is None for cell in row):
            continue

        if len(row) < 3:
            errors.append(f"Row {idx}: incomplete row")
            continue

        raw_date, raw_category, raw_sum = row[:3]

        try:
            parsed_date = raw_date if isinstance(raw_date, date) else datetime.strptime(str(raw_date),
                                                                                        "%Y-%m-%d").date()
        except Exception:
            errors.append(f"Row {idx}: invalid date format")
            continue

        if parsed_date.day != 1:
            errors.append(f"Row {idx}: date must be the first day of the month")
            continue

        if raw_sum is None:
            errors.append(f"Row {idx}: sum cannot be null")
            continue

        try:
            amount = Decimal(str(raw_sum))
        except Exception:
            errors.append(f"Row {idx}: invalid sum value")
            continue

        result = await db.execute(select(Dictionary).where(Dictionary.name == str(raw_category).strip()))
        category = result.scalar_one_or_none()

        if not category:
            errors.append(f"Row {idx}: category not found")
            continue

        existing = await db.execute(
            select(Plan).where(Plan.period == parsed_date, Plan.category_id == category.id)
        )
        existing_plan = existing.scalar_one_or_none()

        if existing_plan:
            print(f"Plan already exists: Period: {existing_plan.period}, Category ID: {existing_plan.category_id}")
            continue

        plans_to_insert.append(Plan(period=parsed_date, sum=amount, category_id=category.id))

    if errors:
        raise HTTPException(status_code=400, detail=errors)

    result = await db.execute(text("SELECT MAX(id) FROM plans"))
    max_id = result.scalar()
    if max_id is None:
        max_id = 1
    await db.execute(text(f"SELECT setval('plans_id_seq', {max_id}, true)"))
    await db.commit()

    try:
        db.add_all(plans_to_insert)
        await db.commit()
    except Exception as e:
        await db.rollback()
        raise HTTPException(status_code=500, detail=f"Database error: {str(e)}")

    return {"detail": f"{len(plans_to_insert)} plans inserted"}

